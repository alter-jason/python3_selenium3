import math
from openpyxl import load_workbook

# addr = r"C:\Users\ASUS\Desktop\phone_positoin.xlsx"
#
# wb = load_workbook(addr)
#
# ws = wb.get_sheet_by_name("phone")  #获取指定单元表
#
# col = []
# pis = []
# print(ws.iter_cols())
# for i  in ws.iter_cols():
# 	# print(i)
# 	col.append(i)
#
#
# num = 0
#
# for i  in range(200001):  #
# 	if i ==0:
# 		pass
# 	else:
# 		pis.append(col[3][i].value)#获取坐标值
# print(pis.count(1))
def get_area(latitude, longitude, dis):
    """
    确定查询经纬度范围
    :param latitude:中心纬度
    :param longitude:中心经度
    :param dis:半径
    :return:(minlat, maxlat, minlng, maxlng)
    """
    r = 6371.137 #地球半径 单位千米
    dlng = 2 * math.asin(math.sin(dis / (2 * r)) / math.cos(latitude * math.pi / 180))
    dlng = dlng * 180 / math.pi

    dlat = dis / r
    dlat = dlat * 180 / math.pi

    minlat = format((latitude - dlat),'.6f')

    maxlat = format((latitude +dlat),'.6f')

    minlng = format((longitude - dlng ),'.6f')

    maxlng = format((longitude +dlng ),'.6f')

    return minlat, maxlat, minlng, maxlng


def distance(lon1, lat1, lon2, lat2):  # 经度1，纬度1，经度2，纬度2 （十进制度数）
    """
    根据经纬度计算距离
    :param lon1: 点1经度
    :param lat1: 点1纬度
    :param lon2: 点2经度
    :param lat2: 点2纬度
    :return:distance
    """
    # 将十进制度数转化为弧度
    lon1, lat1, lon2, lat2 = map(math.radians, [float(lon1), float(lat1), float(lon2), float(lat2)])

    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    c = 2 * math.asin(math.sqrt(a))
    r = 6371.137  # 地球平均半径，单位为公里
    return float('%.2f' % (c * r))
