
'''将高德地图获取的经纬度通过百度地图api转换成百度经纬度'''

import  requests
from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.utils import get_column_letter
import math
from test_p.Base import get_area





ss = get_area(latitude=106.680831,longitude=26.62587,dis=500)


addr = r"C:\Users\ASUS\Desktop\车辆数据.xlsx"

wb = load_workbook(addr)

ws = wb.get_sheet_by_name("Sheet1")  #获取指定单元表

col = []  #行数
pis = []
print(ws.iter_cols())
for i  in ws.iter_cols():
	# print(i)
	col.append(i)


num = 0

for i  in range(1000):  #
	if i ==0:
		pis.append(col[2][i].value)  # 获取坐标值
	else:
		pis.append(col[2][i].value)#获取坐标值
print(pis)
print(type(pis[0]))
sheet = wb['Sheet1']

n = 1
for i  in range(len(pis)):
	url = "http://api.map.baidu.com/geoconv/v1/?coords="+pis[i]+"&from=3&to=5&ak=Dd2HGUBpRHkaI9iLe2skgHmAfRTbGima"
	rq =requests.get(url)

	lnati = str(rq.json()['result'][0]['x'])+','+str(rq.json()['result'][0]['y'])
	ws['D'+str(n)] = lnati
	n +=1
wb.save(addr)
wb.close()