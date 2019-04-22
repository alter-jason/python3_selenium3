
import datetime
from random import choice
import random
from time import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


addr = r"C:\Users\ASUS\Desktop\phone_positoin.xlsx"

wb = load_workbook(addr)

ws = wb.get_sheet_by_name("phone")
#天数列表
TIME1 = ['2019-2-06','2019-2-07','2019-2-08','2019-2-09','2019-2-10','2019-2-11'
        , '2019-2-12','2019-2-13','2019-2-14','2019-2-15','2019-2-16','2019-2-17'
        ,'2019-2-18','2019-2-19','2019-2-20','2019-2-21','2019-2-22','2019-2-23','2019-2-24']
TIME2 = []
#黑车电话生成
heiphone = []
for i  in range(649,999):
    if i%10 ==0:
        heiphone.append("18096075"+str(i))
    elif i%10 ==1:
        heiphone.append("13381865"+str(i))
    elif i%10 ==2:
        heiphone.append("15140851"+str(i))
    elif i%10 ==3:
        heiphone.append("13980865"+str(i))
    elif i%10 ==4:
        heiphone.append("17780951"+str(i))
    elif i%10 ==5:
        heiphone.append("18180955"+str(i))
    else:
        heiphone.append("15680859"+str(i))
#取不重复的时间点
while True:
    ss = choice(TIME1) + ' ' + str(choice(range(7, 24))) + ':' + str(choice(range(1, 60)))
    if ss not in TIME2:
        TIME2.append(ss)
    else:
        continue
    if len(TIME2) == 400:
        break

	# print(len(TIME))
time1 = random.sample(TIME2,400)
for  x  in heiphone:
    for i in range(400):
        if i<=280:
            TIME = time1[i]
            phome =x
            positoin1 = "26.62"+str(choice(range(3441,8592)))+','+"106.6"+str(choice(range(76334,85328)))
            ws.append([phome, TIME, positoin1])
        else:
            TIME = time1[i]
            phome = x
            positoin1 = "26.6" + str(choice(range(33441, 77792))) + ',' + "106." + str(choice(range(595445, 706481)))
            ws.append([phome, TIME, positoin1])
wb.save(addr)


#
# addr = r"C:\Users\ASUS\Desktop\phone_positoin.xlsx"
#
# wb = load_workbook(addr)
#
# ws = wb.get_sheet_by_name("phone")
# #天数列表
# TIME1 = ['2019-2-06','2019-2-07','2019-2-08','2019-2-09','2019-2-10','2019-2-11'
#         , '2019-2-12','2019-2-13','2019-2-14','2019-2-15','2019-2-16','2019-2-17'
#         ,'2019-2-18','2019-2-19','2019-2-20','2019-2-21','2019-2-22','2019-2-22','2019-2-23','2019-2-24']
# TIME2 = []
# #正常车电话生成
# heiphone = []
# for i  in range(300,450):
#     if i%10 ==0:
#         heiphone.append("18096075"+str(i))
#     elif i%10 ==1:
#         heiphone.append("13381865"+str(i))
#     elif i%10 ==2:
#         heiphone.append("15140851"+str(i))
#     elif i%10 ==3:
#         heiphone.append("13980865"+str(i))
#     elif i%10 ==4:
#         heiphone.append("17780951"+str(i))
#     elif i%10 ==5:
#         heiphone.append("18180955"+str(i))
#     else:
#         heiphone.append("15680859" + str(i))
# #取不重复的时间点
# while True:
#     ss = choice(TIME1) + ' ' + str(choice(range(7, 24))) + ':' + str(choice(range(1, 60)))
#     if ss not in TIME2:
#         TIME2.append(ss)
#     else:
#         continue
#     if len(TIME2) == 400:
#         break
# time1 = random.sample(TIME2,400)
# for  x  in heiphone:
#     for i in range(400):
#         if i >=393:
#             TIME = time1[i]
#             phome =x
#             positoin1 = "26.62"+str(choice(range(3441,8592)))+','+"106.6"+str(choice(range(79123,85648)))
#             ws.append([phome, TIME, positoin1])
#         else:
#             TIME = time1[i]
#             phome = x
#             positoin1 = "26.6" + str(choice(range(33441, 77792))) + ',' + "106." + str(choice(range(595445, 706481)))
#             ws.append([phome, TIME, positoin1])
# wb.save(addr)