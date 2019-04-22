from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.utils import get_column_letter
import math
from test_p.Base import get_area





ss = get_area(latitude=106.680831,longitude=26.62587,dis=500)


addr = r"C:\Users\ASUS\Desktop\aa.xlsx"

wb = load_workbook(addr)

ws = wb.get_sheet_by_name("phone")  #获取指定单元表

col = []  #行数
pis = []
print(ws.iter_cols())
for i  in ws.iter_cols():
	# print(i)
	col.append(i)


num = 0

for i  in range(7000):  #
	if i ==0:
		pis.append(col[2][i].value)  # 获取坐标值
	else:
		pis.append(col[2][i].value)#获取坐标值
print(pis)
print(type(pis[0]))
#
# num = []
# for i  in range(len(pis)):
# 	n1 = pis[i].split(',')
# 	print(n1)
# 	print(float( n1[0]))
# 	if float(ss[3])<=float(n1[0]) <= float(ss[2]):
# 		if float(ss[0])<=float(n1[1]) <= float(ss[1]):
# 			num.append(1)
# 			# print( ' postion is in')
# 		else:
# 			num.append(0)
# 			# print("not  in")
# 	else:
# 		num.append(0)
# 		# print('not in')
# i=2
# for ex  in num:
# 	ws['D'+str(i)] = ex #d列循环写入数据
# 	i += 1
# wb.save(addr)